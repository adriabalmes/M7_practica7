#importamos la libreria  load_workbook
import openpyxl
from openpyxl import load_workbook

#creamos la funcion
from openpyxl.utils.dataframe import dataframe

def function_createxcel():
    #Creamos el archivo
    wb = load_workbook("file.xlsx")
    sheet = wb.worksheets[0]
    #Rellenamos los datos
    name=["Guillem", "Adrian", "Sebastian"]
    surname=["Jimenez","Berruezo","Casco"]
    age=["30","34","19"]

    sheet.cell(row=1, column=1).value="Name"
    sheet.cell(row=1, column=2).value="Surname"
    sheet.cell(row=1, column=3).value="Age"

    j=2
    for i in range(0,5):
        sheet.cell(row=j, column=1).value=name[i]
        sheet.cell(row=j, column=2).value=surname[i]
        sheet.cell(row=j, column=3).value=age[i]
        j+=1

    wb.save("excel.xlsx")

def function_readexcel():
# Definimos la varible dataframe
    dataframe = openpyxl.load_workbook("file.xlsx")

# Definimos la variable para leer
dataframe1 = dataframe.active

# Bucle para leer los datos
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)

function_createxcel()
function_readexcel()
